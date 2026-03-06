"""File read/write/inspect tools — sandboxed to project root, blacklist + node access enforced."""

from __future__ import annotations

import json as _json
import struct
from pathlib import Path

from langchain_core.tools import tool

from baymax.permissions import assert_not_blacklisted, assert_node_access


def _resolve_safe(project_root: str, file_path: str) -> Path:
    """Resolve a path and ensure it stays within project root."""
    root = Path(project_root).resolve()
    target = (root / file_path).resolve()
    if not str(target).startswith(str(root)):
        raise ValueError(f"Path escapes project root: {file_path}")
    return target


# ── Format-specific inspectors ────────────────────────────────

def _inspect_csv(path: Path, head: int) -> str:
    import pandas as pd
    df_sample = pd.read_csv(path, nrows=head)
    with open(path) as f:
        row_count = sum(1 for _ in f) - 1
    lines = [
        f"CSV: {path.name}",
        f"Shape: {row_count} rows x {len(df_sample.columns)} columns",
        f"Columns: {list(df_sample.columns)}",
        f"Dtypes:\n{df_sample.dtypes.to_string()}",
        f"\nFirst {head} rows:\n{df_sample.head(head).to_string()}",
    ]
    return "\n".join(lines)


def _inspect_json(path: Path) -> str:
    raw = path.read_text()
    try:
        data = _json.loads(raw)
    except _json.JSONDecodeError as e:
        return f"JSON: {path.name} (parse error: {e})\nFirst 500 chars:\n{raw[:500]}"

    if isinstance(data, list):
        summary = f"JSON array with {len(data)} items"
        if data:
            first = data[0]
            if isinstance(first, dict):
                summary += f"\nItem keys: {list(first.keys())}"
            sample = _json.dumps(data[:3], indent=2, default=str)
            if len(sample) > 2000:
                sample = sample[:2000] + "\n... (truncated)"
            summary += f"\nFirst 3 items:\n{sample}"
    elif isinstance(data, dict):
        summary = f"JSON object with {len(data)} top-level keys"
        summary += f"\nKeys: {list(data.keys())}"
        sample = _json.dumps(data, indent=2, default=str)
        if len(sample) > 2000:
            sample = sample[:2000] + "\n... (truncated)"
        summary += f"\nContent:\n{sample}"
    else:
        summary = f"JSON scalar: {type(data).__name__} = {str(data)[:500]}"

    return f"JSON: {path.name}\n{summary}"


def _inspect_image(path: Path) -> str:
    size_bytes = path.stat().st_size
    ext = path.suffix.lower()
    dims = "unknown"
    try:
        with open(path, "rb") as f:
            header = f.read(32)
            if ext == ".png" and header[:8] == b"\x89PNG\r\n\x1a\n":
                w, h = struct.unpack(">II", header[16:24])
                dims = f"{w} x {h}"
            elif ext in (".jpg", ".jpeg"):
                f.seek(0)
                data = f.read()
                idx = 2
                while idx < len(data) - 1:
                    marker = data[idx:idx + 2]
                    if marker[0] != 0xFF:
                        break
                    if marker[1] in (0xC0, 0xC2):
                        h, w = struct.unpack(">HH", data[idx + 5:idx + 9])
                        dims = f"{w} x {h}"
                        break
                    length = struct.unpack(">H", data[idx + 2:idx + 4])[0]
                    idx += 2 + length
            elif ext == ".gif" and header[:6] in (b"GIF87a", b"GIF89a"):
                w, h = struct.unpack("<HH", header[6:10])
                dims = f"{w} x {h}"
            elif ext == ".webp" and header[:4] == b"RIFF" and header[8:12] == b"WEBP":
                dims = "(webp — dimensions require full parse)"
    except Exception:
        pass

    size_str = (
        f"{size_bytes / 1_048_576:.1f} MB" if size_bytes > 1_048_576
        else f"{size_bytes / 1024:.1f} KB"
    )
    return f"Image: {path.name}\nFormat: {ext.lstrip('.')}\nDimensions: {dims}\nSize: {size_str}"


def _inspect_excel(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = [f"Excel: {path.name}", f"Sheets: {wb.sheetnames}"]
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(max_row=1, values_only=True))
            cols = list(rows[0]) if rows else []
            lines.append(f"  {name}: columns={cols}, rows~={ws.max_row}")
        wb.close()
        return "\n".join(lines)
    except ImportError:
        return f"Excel: {path.name}\n[Cannot inspect — openpyxl not installed. Use shell: python3 -c \"import openpyxl; ...\"]"
    except Exception as e:
        return f"Excel: {path.name}\n[error] {e}"


def _inspect_pdf(path: Path) -> str:
    for reader_fn in [_pdf_via_pymupdf, _pdf_via_pdfplumber]:
        result = reader_fn(path)
        if result:
            return result
    size_bytes = path.stat().st_size
    size_str = f"{size_bytes / 1_048_576:.1f} MB" if size_bytes > 1_048_576 else f"{size_bytes / 1024:.1f} KB"
    return (
        f"PDF: {path.name}\nSize: {size_str}\n"
        "[Cannot inspect — no PDF library available. "
        "Use shell: python3 -c \"import subprocess; subprocess.run(['pdftotext', ...])\"]"
    )


def _pdf_via_pymupdf(path: Path) -> str | None:
    try:
        import fitz  # pymupdf
        doc = fitz.open(str(path))
        pages = doc.page_count
        text = ""
        for i in range(min(2, pages)):
            text += doc[i].get_text()
        doc.close()
        if len(text) > 3000:
            text = text[:3000] + "\n... (truncated)"
        return f"PDF: {path.name}\nPages: {pages}\n\nText (first 2 pages):\n{text}"
    except ImportError:
        return None


def _pdf_via_pdfplumber(path: Path) -> str | None:
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            pages = len(pdf.pages)
            text = ""
            for page in pdf.pages[:2]:
                text += (page.extract_text() or "") + "\n"
        if len(text) > 3000:
            text = text[:3000] + "\n... (truncated)"
        return f"PDF: {path.name}\nPages: {pages}\n\nText (first 2 pages):\n{text}"
    except ImportError:
        return None


def _inspect_text(path: Path, max_lines: int = 200) -> str:
    try:
        text = path.read_text()
    except UnicodeDecodeError:
        size_bytes = path.stat().st_size
        with open(path, "rb") as f:
            hex_preview = f.read(32).hex(" ")
        return (
            f"Binary file: {path.name}\n"
            f"Size: {size_bytes} bytes\n"
            f"First 32 bytes (hex): {hex_preview}\n"
            "[Cannot read as text. Use shell to probe: file <path>, xxd <path> | head]"
        )

    lines = text.splitlines()
    total = len(lines)
    if total > max_lines:
        preview = "\n".join(lines[:max_lines])
        return f"Text: {path.name} ({total} lines, showing first {max_lines})\n\n{preview}\n... ({total - max_lines} more lines)"
    return f"Text: {path.name} ({total} lines)\n\n{text}"


# ── Tool factories ────────────────────────────────────────────

_FORMAT_MAP = {
    ".csv": "csv", ".tsv": "csv",
    ".json": "json", ".jsonl": "json",
    ".pdf": "pdf",
    ".png": "image", ".jpg": "image", ".jpeg": "image", ".gif": "image", ".webp": "image",
    ".xlsx": "excel", ".xls": "excel",
}


def create_file_inspect_tool(project_root: str, node_name: str = "planner"):
    @tool
    def file_inspect(file_path: str, head: int = 5) -> str:
        """Inspect a file intelligently. Auto-detects format (CSV, JSON, PDF, image, Excel, text)
        and returns a useful summary instead of raw content. Use this to understand file structure
        and schema without dumping the entire file into context.

        Args:
            file_path: Path relative to the project root.
            head: Number of sample rows for CSV/TSV (default 5).
        """
        assert_not_blacklisted(file_path)
        assert_node_access(file_path, node_name)
        target = _resolve_safe(project_root, file_path)
        if not target.exists():
            return f"[error] File not found: {file_path}"

        ext = target.suffix.lower()
        fmt = _FORMAT_MAP.get(ext, "text")

        try:
            if fmt == "csv":
                return _inspect_csv(target, head)
            elif fmt == "json":
                return _inspect_json(target)
            elif fmt == "pdf":
                return _inspect_pdf(target)
            elif fmt == "image":
                return _inspect_image(target)
            elif fmt == "excel":
                return _inspect_excel(target)
            else:
                return _inspect_text(target)
        except Exception as e:
            return f"[error] Failed to inspect {file_path}: {e}"

    return file_inspect


def create_file_read_tool(project_root: str, node_name: str = "executor"):
    @tool
    def file_read(file_path: str) -> str:
        """Read a file relative to the project root. Returns the raw file contents."""
        assert_not_blacklisted(file_path)
        assert_node_access(file_path, node_name)
        target = _resolve_safe(project_root, file_path)
        if not target.exists():
            return f"[error] File not found: {file_path}"
        return target.read_text()

    return file_read


def create_file_write_tool(project_root: str, node_name: str = "executor"):
    @tool
    def file_write(file_path: str, content: str) -> str:
        """Write content to a file relative to the project root. Creates parent directories if needed."""
        assert_not_blacklisted(file_path)
        assert_node_access(file_path, node_name)
        target = _resolve_safe(project_root, file_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(content)
        return f"Wrote {len(content)} bytes to {file_path}"

    return file_write
